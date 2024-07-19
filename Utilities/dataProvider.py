import openpyxl



def get_data(sheet_name):
    """
    Reads data from an Excel sheet and returns a list of lists,
    ensuring numeric values are handled correctly.

    Args:
        sheet_name (str): The name of the sheet to read data from.

    Returns:
        list: A list of lists containing the data from the Excel sheet.
    """

    try:
        workbook = openpyxl.load_workbook("..//excel//testdata1.xlsx")  # Adjust path if needed
        sheet = workbook[sheet_name]
        total_rows = sheet.max_row
        total_cols = sheet.max_column

        data_list = []  # Use a more descriptive variable name

        for row_index in range(2, total_rows + 1):  # Use meaningful variable names
            row_data = []
            for col_index in range(1, total_cols + 1):
                cell = sheet.cell(row=row_index, column=col_index)

                # Handle potential errors and data types
                try:
                    value = cell.value
                except Exception as e:
                    print(f"Error reading cell {row_index},{col_index}: {e}")
                    value = None  # Or handle differently as needed

                if isinstance(value, (int, float)):  # Check for numeric types
                    row_data.append(value)
                else:
                    row_data.append(value)  # Include non-numeric values

            data_list.append(row_data)

        return data_list

    except FileNotFoundError as e:
        print(f"Error: Excel file '..//excel//testdata1.xlsx' not found: {e}")
        return None  # Or handle differently as needed
